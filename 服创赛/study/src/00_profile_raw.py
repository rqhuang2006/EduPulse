from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from study_common import DOCS_DIR, RAW_DIR, ensure_dirs, file_registry


def profile_workbook(path: Path) -> tuple[list[dict], list[dict]]:
    sheet_rows: list[dict] = []
    field_rows: list[dict] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    excel = pd.ExcelFile(path)

    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        sample = pd.read_excel(path, sheet_name=sheet, nrows=1000)
        sample.columns = [str(c).strip() for c in sample.columns]
        sheet_rows.append(
            {
                "SOURCE_FILE": path.name,
                "SHEET_NAME": sheet,
                "ROWS": max(ws.max_row - 1, 0),
                "COLUMNS": ws.max_column,
                "PROFILED_SAMPLE_ROWS": len(sample),
            }
        )
        for col in sample.columns:
            values = sample[col].dropna().astype(str).head(5).tolist()
            field_rows.append(
                {
                    "SOURCE_FILE": path.name,
                    "SHEET_NAME": sheet,
                    "RAW_FIELD": col,
                    "SAMPLE_NULL_RATE": float(sample[col].isna().mean()) if len(sample) else None,
                    "SAMPLE_VALUES": " | ".join(values),
                }
            )

    excel.close()
    workbook.close()
    return sheet_rows, field_rows


def main() -> None:
    ensure_dirs()
    registry = file_registry()
    dataset_by_file = {v["raw_file"]: k for k, v in registry.items()}
    all_sheets: list[dict] = []
    all_fields: list[dict] = []

    for path in sorted(RAW_DIR.glob("*.xls*")):
        sheet_rows, field_rows = profile_workbook(path)
        dataset = dataset_by_file.get(path.name, "")
        for row in sheet_rows:
            row["DATASET"] = dataset
        for row in field_rows:
            row["DATASET"] = dataset
        all_sheets.extend(sheet_rows)
        all_fields.extend(field_rows)

    output = DOCS_DIR / "raw_inventory.xlsx"
    with pd.ExcelWriter(output) as writer:
        pd.DataFrame(all_sheets).to_excel(writer, sheet_name="sheets", index=False)
        pd.DataFrame(all_fields).to_excel(writer, sheet_name="fields", index=False)

    print(f"raw inventory written: {output}")


if __name__ == "__main__":
    main()

